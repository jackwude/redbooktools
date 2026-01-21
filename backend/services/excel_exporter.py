"""
Excel 导出服务
将舆情分析报告导出为 Excel 文件
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel 导出器"""
    
    def __init__(self):
        self.wb = None
        
    async def export_analysis_report(self, report_data: dict) -> BytesIO:
        """
        导出分析报告为 Excel 文件
        
        Args:
            report_data: 分析报告数据字典
            
        Returns:
            BytesIO: Excel 文件的字节流
        """
        try:
            self.wb = Workbook()
            
            # 移除默认的工作表
            self.wb.remove(self.wb.active)
            
            # 创建多个工作表
            self._create_summary_sheet(report_data)
            self._create_posts_sheet(report_data)
            self._create_sentiment_sheet(report_data)
            self._create_keywords_sheet(report_data)
            
            if report_data.get('risk_alerts'):
                self._create_alerts_sheet(report_data)
            
            # 保存到内存
            excel_file = BytesIO()
            self.wb.save(excel_file)
            excel_file.seek(0)
            
            logger.info("Excel 报告导出成功")
            return excel_file
            
        except Exception as e:
            logger.error(f"Excel 导出失败: {e}", exc_info=True)
            raise
    
    def _create_summary_sheet(self, report_data: dict):
        """创建概览工作表"""
        ws = self.wb.create_sheet("📊 分析概览")
        
        # 标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='FF2442', end_color='FF2442', fill_type='solid')
        
        # 标题
        ws['A1'] = '小红书舆情分析报告'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # 基本信息
        row = 3
        info_items = [
            ('报告 ID', report_data.get('analysis_id', 'N/A')),
            ('生成时间', datetime.fromisoformat(report_data.get('created_at', datetime.now().isoformat())).strftime('%Y-%m-%d %H:%M:%S')),
            ('搜索关键词', report_data.get('search_keyword', '未指定')),
            ('识别帖子数', report_data.get('total_posts', 0)),
        ]
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='微软雅黑', bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # 情感分布
        row += 1
        ws[f'A{row}'] = '情感分布'
        ws[f'A{row}'].font = Font(name='微软雅黑', size=12, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        dist = report_data.get('sentiment_distribution', {})
        sentiment_data = [
            ('正面', dist.get('positive_count', 0), f"{dist.get('positive_ratio', 0) * 100:.1f}%", '22C55E'),
            ('中性', dist.get('neutral_count', 0), f"{dist.get('neutral_ratio', 0) * 100:.1f}%", '3B82F6'),
            ('负面', dist.get('negative_count', 0), f"{dist.get('negative_ratio', 0) * 100:.1f}%", 'EF4444'),
        ]
        
        ws[f'A{row}'] = '情感'
        ws[f'B{row}'] = '数量'
        ws[f'C{row}'] = '占比'
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(name='微软雅黑', bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        row += 1
        for sentiment, count, ratio, color in sentiment_data:
            ws[f'A{row}'] = sentiment
            ws[f'B{row}'] = count
            ws[f'C{row}'] = ratio
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[f'A{row}'].font = Font(color='FFFFFF', bold=True)
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_posts_sheet(self, report_data: dict):
        """创建帖子详情工作表"""
        ws = self.wb.create_sheet("📝 帖子详情")
        
        # 表头
        headers = ['序号', '标题', '内容摘要', '情感', '点赞数', '评论数', '关键词']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        posts = report_data.get('posts', [])
        sentiment_colors = {
            'positive': '22C55E',
            'neutral': '3B82F6',
            'negative': 'EF4444',
        }
        sentiment_labels = {
            'positive': '正面',
            'neutral': '中性',
            'negative': '负面',
        }
        
        for idx, post in enumerate(posts, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=post.get('title', ''))
            ws.cell(row=row, column=3, value=post.get('content', '')[:100] + '...' if post.get('content') else '')
            
            # 情感标签
            sentiment = post.get('sentiment', 'neutral')
            sentiment_cell = ws.cell(row=row, column=4, value=sentiment_labels.get(sentiment, '中性'))
            sentiment_cell.fill = PatternFill(
                start_color=sentiment_colors.get(sentiment, '3B82F6'),
                end_color=sentiment_colors.get(sentiment, '3B82F6'),
                fill_type='solid'
            )
            sentiment_cell.font = Font(color='FFFFFF', bold=True)
            sentiment_cell.alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=5, value=post.get('likes', 0))
            ws.cell(row=row, column=6, value=post.get('comments', 0))
            ws.cell(row=row, column=7, value=', '.join(post.get('keywords', [])[:5]))
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def _create_sentiment_sheet(self, report_data: dict):
        """创建情感分析工作表"""
        ws = self.wb.create_sheet("📈 情感分析")
        
        # 标题
        ws['A1'] = '情感分布统计'
        ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # 表头
        headers = ['情感类型', '数量', '占比', '占比（%）', '趋势']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(name='微软雅黑', bold=True)
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        dist = report_data.get('sentiment_distribution', {})
        total = report_data.get('total_posts', 1)
        
        data = [
            ('正面 😊', dist.get('positive_count', 0), dist.get('positive_ratio', 0), '22C55E'),
            ('中性 😐', dist.get('neutral_count', 0), dist.get('neutral_ratio', 0), '3B82F6'),
            ('负面 😞', dist.get('negative_count', 0), dist.get('negative_ratio', 0), 'EF4444'),
        ]
        
        row = 3
        for sentiment, count, ratio, color in data:
            ws.cell(row=row, column=1, value=sentiment).font = Font(bold=True)
            ws.cell(row=row, column=2, value=count).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{ratio:.2%}").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=ratio * 100).alignment = Alignment(horizontal='center')
            
            # 趋势条
            trend_cell = ws.cell(row=row, column=5)
            trend_cell.value = '█' * int(ratio * 20)
            trend_cell.font = Font(color=color)
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
    
    def _create_keywords_sheet(self, report_data: dict):
        """创建关键词工作表"""
        ws = self.wb.create_sheet("🔥 热门关键词")
        
        # 表头
        headers = ['排名', '关键词', '出现次数', '情感倾向']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        keywords = report_data.get('top_keywords', [])
        sentiment_colors = {
            'positive': '22C55E',
            'neutral': '3B82F6',
            'negative': 'EF4444',
        }
        sentiment_labels = {
            'positive': '正面',
            'neutral': '中性',
            'negative': '负面',
        }
        
        for idx, kw in enumerate(keywords, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=kw.get('word', ''))
            ws.cell(row=row, column=3, value=kw.get('count', 0)).alignment = Alignment(horizontal='center')
            
            sentiment = kw.get('sentiment', 'neutral')
            sentiment_cell = ws.cell(row=row, column=4, value=sentiment_labels.get(sentiment, '中性'))
            sentiment_cell.fill = PatternFill(
                start_color=sentiment_colors.get(sentiment, '3B82F6'),
                end_color=sentiment_colors.get(sentiment, '3B82F6'),
                fill_type='solid'
            )
            sentiment_cell.font = Font(color='FFFFFF', bold=True)
            sentiment_cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def _create_alerts_sheet(self, report_data: dict):
        """创建风险预警工作表"""
        ws = self.wb.create_sheet("⚠️ 风险预警")
        
        # 表头
        headers = ['风险等级', '预警描述']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        alerts = report_data.get('risk_alerts', [])
        level_colors = {
            'high': 'EF4444',
            'medium': 'F59E0B',
            'low': '3B82F6',
        }
        level_labels = {
            'high': '🔴 高风险',
            'medium': '🟡 中风险',
            'low': '🔵 低风险',
        }
        
        for idx, alert in enumerate(alerts, 1):
            row = idx + 1
            level = alert.get('level', 'low')
            
            level_cell = ws.cell(row=row, column=1, value=level_labels.get(level, '🔵 低风险'))
            level_cell.fill = PatternFill(
                start_color=level_colors.get(level, '3B82F6'),
                end_color=level_colors.get(level, '3B82F6'),
                fill_type='solid'
            )
            level_cell.font = Font(color='FFFFFF', bold=True)
            level_cell.alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=2, value=alert.get('description', ''))
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80


def get_excel_exporter() -> ExcelExporter:
    """获取 Excel 导出器实例"""
    return ExcelExporter()
